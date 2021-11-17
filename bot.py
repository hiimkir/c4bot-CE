import os
import openpyxl
import connect4
import asyncio
from pathlib import Path
from dotenv import load_dotenv
import discord
from discord.ext import commands
from discord.utils import get
from discord import FFmpegPCMAudio
from yandex_music import Client, TrackShort

load_dotenv()
bot = commands.Bot(command_prefix=':')
jukebox = Client.from_credentials(os.getenv('LOGIN'), os.getenv('PASS'))
Path('./YMcache/').mkdir(parents=True, exist_ok=True)


@bot.command(help='Let the music play')
async def play(ctx, *args):
    voice = get(bot.voice_clients, guild=ctx.guild)
    channel = ctx.message.author.voice

    if not channel:
        await ctx.send('Connect to voice dumbass')
        return
    channel = channel.channel

    if not voice:
        await channel.connect()
        voice = get(bot.voice_clients, guild=ctx.guild)

    if channel != voice:
        await voice.move_to(channel)
        voice = get(bot.voice_clients, guild=ctx.guild)

    arg = None
    forcepush = False
    skip = 0
    for i in args:
        if i[0] == '-':
            skip += 1
            if i == '-f':
                forcepush = True
            else:
                arg = i[1::]
        elif skip >= 0:
            break

    if voice.is_playing() and forcepush:
        voice.stop()

    if not voice.is_playing():
        text = ' '.join(args[skip::])

        if not arg:
            arg = 'all'
            text = text.split('/')
            if len(text) > 1:
                for word in text:
                    if word == 'music.yandex.ru':
                        arg = 'link'
                        break
            text = '/'.join(text)

        try:
            if arg == 'link':
                int(text.split('/')[-1])
                parsed_id = text.split('/')[-1]
                try:
                    int(text.split('/')[-3])
                    parsed_id += ':' + text.split('/')[-3]
    
                    tracks = jukebox.tracks(track_ids=[parsed_id])
                except Exception:
                    setlist = jukebox.albums_with_tracks(parsed_id)
                    tracks = []
                    for volume in setlist.volumes:
                        tracks += volume
    
            elif arg == 'loves':
                tracks = jukebox.users_likes_tracks()
    
            elif arg == 'likes':
                tracks = ['609676:14599232']
    
            else:
                query = jukebox.search(text=text, type_=arg)
                if arg == 'track':
                    setlist = query.tracks.results[0]
                elif arg == 'album':
                    setlist = query.albums.results[0]
                elif arg == 'artist':
                    setlist = query.artists.results[0]
                elif arg == 'playlist':
                    setlist = query.playlists.results[0]
                else:
                    setlist = query.best.result
                    arg = query.best.type
    
                if arg == 'track':
                    tracks = [setlist]
                elif arg == 'album':
                    setlist = setlist.with_tracks()
                    tracks = []
                    for volume in setlist.volumes:
                        tracks += volume
                elif arg == 'artist':
                    tracks = setlist.get_tracks()
                elif arg == 'playlist':
                    tracks = setlist.fetch_tracks()
                else:
                    await ctx.send('Nothing found')
                    return

        except Exception as woopsie:
            await ctx.send(f"I don't get it ||{woopsie}||")
            await voice.disconnect()
            return

        for track in tracks:
            if type(track) == TrackShort:
                track = track.fetch_track()
            file_path = f'./YMcache/{track.id}.aac'  # track should be deleted after some time
            track_time = track.duration_ms // 1000
            time_second = f'0{track_time % 60}' if track_time % 60 < 10 else str(track_time % 60)

            if not os.path.exists(file_path):
                try:
                    track.download(file_path, 'aac', 128)
                except Exception:
                    try:
                        track.download(file_path)
                    except Exception as woopsie:
                        await ctx.send(f'I failed because {woopsie}')
                        return

            voice.play(FFmpegPCMAudio(file_path))
            voice.is_playing()

            body = ''
            if len(track.artists_name()) > 1:
                body = 'feat. ' + ' '.join(track.arists_name()[1::])
            elif track.albums[0].short_description:
                body = track.albums[0].short_description
            elif track.albums[0].description:
                body = track.albums[0].description.split('.')[0] + '.'
            output_msg = discord.Embed(title=f'{track.artists[0].name} - {track.title}', color=0xFFDB4E,
                                       url=f'https://music.yandex.ru/album/{track.albums[0].id}/track/{track.id}',
                                       description=body)
            output_msg.set_author(name=f'added by {ctx.author.display_name}', icon_url=ctx.author.avatar_url)
            output_msg.set_thumbnail(url=f'https://{track.cover_uri[:-2]}1000x1000')
            output_msg.set_footer(text=f'{track_time // 60}:{time_second}',
                                  icon_url='https://img.icons8.com/fluency-systems-filled/48/ffffff/time.png')
            await ctx.send(embed=output_msg, delete_after=track_time)
            await asyncio.sleep(track_time + 1)

    else:
            output_msg = discord.Embed(title=f'Queue coming soon!', url='https://github.com/hiimkir/c4bot-CE',
                                       description='just kick the bot', color=0xF54542)
            await ctx.send(embed=output_msg, delete_after=180)
            return


@bot.command(help='There is no room for this bot anymore')
async def gtfo(ctx):
    voice = get(bot.voice_clients, guild=ctx.guild)

    if voice and voice.is_connected():
        if voice.is_playing():
            voice.stop()
        await voice.disconnect()


@bot.command(name='69', help='Funny number')
async def sixty_nine(ctx):
    await ctx.send('Nice!')


@bot.command(name='c4', help='A game for two (enter number 1-7, ping to add opponent)')
async def connect_four(ctx, cont: str):         # use *args
    if not connect4.game_started(ctx.channel.id):
        connect4.new_game(ctx.channel.id)

    r = openpyxl.load_workbook('save.xlsx')
    sheet = r[str(ctx.channel.id)]  # unload xlsx from memory
    p1 = sheet['A2'].value
    p2 = sheet['B2'].value
    message_author = "'" + str(ctx.author.id)

    if message_author == p2:
        await ctx.send('Next!')
        return

    if p1 == "'0":
        if len(cont) > 21 and cont.split()[0][1] == '@':
            if len(cont) > 26 and cont.split()[1] == '-pass':
                p1 = "'" + cont.split()[1][3:-1]
                p2 = message_author
            else:
                p2 = "'" + cont.split()[1][3:-1]
                p1 = message_author
            sheet.cell(column=1, row=2, value=p1)
            sheet.cell(column=2, row=2, value=p2)
            r.save('save.xlsx')
            await ctx.send('Ready player one')
            return
        sheet.cell(column=1, row=2, value=message_author)
        r.save('save.xlsx')

    elif message_author != p1:
        await ctx.send('Wait a minute, who are you?')
        return

    player_input = int(cont.split()[0]) - 1
    if player_input == 68:
        await ctx.send('Nice!')
        return
    if player_input > 6 or player_input < 0:
        await ctx.end('Try another number 1-7')
        return

    response = connect4.turn(player_input, ctx.channel.id)

    if not response:
        await ctx.send('Column is full')
        return
    await ctx.send(response)

    connect4.end_turn(ctx.channel.id)


@bot.command(help='Clear the table')
async def clear(ctx):
    r = openpyxl.load_workbook('save.xlsx')

    for i in r.worksheets:
        if str(i)[12:-2] == str(ctx.channel.id):
            r.remove(i)
            r.save('save.xlsx')
            await ctx.send('Territory purged')
            return
    await ctx.send('Wow, so empty')


@bot.event
async def on_ready():
    print(f'logged in as {bot.user.name}')


bot.run(os.getenv('TOKEN'))
