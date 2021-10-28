import openpyxl


def end_turn(num):
    r = openpyxl.load_workbook('save.xlsx')
    sheet = r[str(num)]
    plr = (sheet['C2'].value + 1) % 2
    p1 = sheet['A2'].value
    p2 = sheet['B2'].value
    sheet.cell(column=3, row=2, value=plr)
    sheet.cell(column=2, row=2, value=p1)
    sheet.cell(column=1, row=2, value=p2)
    r.save('save.xlsx')
    return


def turn(col, num):
    r = openpyxl.load_workbook('save.xlsx')
    sheet = r[str(num)]
    plr = sheet['C2'].value
    grid = []
    for i in range(3, 10):
        gridline = []
        for j in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            gridline.append(sheet[j + str(i)].value)
        grid.append(gridline)
    ln = 6
    while grid[ln][col] != ':white_circle:':
        ln -= 1
        if ln == -1:
            return False
    if plr == 0:
        grid[ln][col] = ':red_circle:'
    else:
        grid[ln][col] = ':blue_circle:'
    sheet.cell(column=(col + 1), row=(ln + 3), value=grid[ln][col])

    r.save('save.xlsx')
    line = ''
    for i in range(7):
        for j in grid[i]:
            line += j
        line += '\n'
    return line


def new_game(num):
    r = openpyxl.load_workbook('save.xlsx')                              # {Workbook}.copy_worksheet()
    new_sheet = r.create_sheet(title=str(num))
    new_sheet.cell(column=1, row=2, value="'0")
    new_sheet.cell(column=2, row=2, value="'0")
    for i in range(3, 8):
        new_sheet.cell(column=i, row=2, value=0)
    for i in range(1, 8):
        for j in range(3, 10):
            new_sheet.cell(column=i, row=j, value=':white_circle:')
    r.save('save.xlsx')
    return


def game_started(num):                                   # game_started and new_game can be merged
    r = openpyxl.load_workbook('save.xlsx')              # may be too slow
    for i in r.worksheets:
        if str(i)[12:-2] == str(num):
            return True
    return False
                                                                       